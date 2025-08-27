import json
import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from pathlib import Path
import openpyxl

def main():
    def browse_file():
        file_path = filedialog.askopenfilename(
            title="Datei auswählen",
            filetypes=[("JSON Dateien", "*.json"), ("Alle Dateien", "*.*")]
        )
        if file_path:
            entry_var.set(file_path)

    def export_artifacts():
        json_path = entry_var.get()
        if not json_path:
            messagebox.showerror("Fehler", "Bitte zuerst eine Datei auswählen!")
            return
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            artifacts = data.get("artifacts", [])
            if not artifacts:
                messagebox.showerror("Fehler", "Keine 'artifacts' im JSON gefunden!")
                return

            # Artefakte-Tabelle: Jede sec_effect als eigene Zeile, inkl. unit_id
            artifact_rows = []
            # Mapping von rid zu unit_id (Artefakt-IDs können Dicts sein)
            unit_id_map = {}
            for mon in data.get("unit_list", []):
                if mon.get("artifacts"):
                    for arti in mon["artifacts"]:
                        # Falls arti ein Dict ist, hole die rid
                        if isinstance(arti, dict):
                            rid = arti.get("rid")
                        else:
                            rid = arti
                        unit_id_map[rid] = mon.get("unit_id")

            for art in artifacts:
                row = {
                    "rid" : art.get("rid"),
                    "occupied_id" : art.get("occupied_id"),
                    "type": art.get("type"),
                    "attribute": art.get("attribute", "type"),
                    "unit_style": art.get("unit_style"),
                    "natural_rank": art.get("natural_rank"),
                    "pri_effect": art.get("pri_effect"),
                    #stat1
                    "stat1":        art.get("sec_effects", [[None]*5]*4)[0][0] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat1roll":    art.get("sec_effects", [[None]*5]*4)[0][1] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat1crafted": art.get("sec_effects", [[None]*5]*4)[0][3] if len(art.get("sec_effects", [])) > 0 and len(art.get("sec_effects", [[None]*5]*4)[0]) > 2 else None,
                    #stat2
                    "stat2":        art.get("sec_effects", [[None]*5]*4)[1][0] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat2roll":    art.get("sec_effects", [[None]*5]*4)[1][1] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat2crafted": art.get("sec_effects", [[None]*5]*4)[1][3] if len(art.get("sec_effects", [])) > 1 and len(art.get("sec_effects", [[None]*5]*4)[1]) > 2 else None,
                    #stat3
                    "stat3":        art.get("sec_effects", [[None]*5]*4)[2][0] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat3roll":    art.get("sec_effects", [[None]*5]*4)[2][1] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat3crafted": art.get("sec_effects", [[None]*5]*4)[2][3] if len(art.get("sec_effects", [])) > 2 and len(art.get("sec_effects", [[None]*5]*4)[2]) > 2 else None,
                    #stat4
                    "stat4":        art.get("sec_effects", [[None]*5]*4)[3][0] if art.get("sec_effects", [[None]*5]*4)[1] else None,
                    "stat4roll":    art.get("sec_effects", [[None]*5]*4)[3][1] if art.get("sec_effects", [[None]*5]*4)[0] else None,
                    "stat4crafted": art.get("sec_effects", [[None]*5]*4)[3][3] if len(art.get("sec_effects", [])) > 3 and len(art.get("sec_effects", [[None]*5]*4)[3]) > 2 else None,

                    "category": art.get("locked"),
                    "extra": str(art.get("extra")),
                    "date_add": art.get("date_add"),
                    "date_mod": art.get("date_mod"),
                }
                artifact_rows.append(row)

            artifact_df = pd.DataFrame(artifact_rows)
            # Monster-Tabelle: Jede Unit-Artefakt-Kombination als eigene Zeile
            monster = data.get("unit_list", [])
            monster_rows = []
            if not monster:
                messagebox.showerror("Fehler", "Keine 'monsterlist' im JSON gefunden!")
                return
            for mon in monster:
                if mon.get("artifacts"):
                    for arti in mon.get("artifacts"): #die Artefakte links und rechts
                        row = {
                            "rid" : arti.get("rid"),
                            "occupied_id" : arti.get("occupied_id"),
                            "type": arti.get("type"),
                            "attribute": arti.get("attribute", "type"),
                            "unit_style": arti.get("unit_style"),
                            "natural_rank": arti.get("natural_rank"),
                            "pri_effect": arti.get("pri_effect"),
                            #stat1
                            "stat1":        arti.get("sec_effects", [[None]*5]*4)[0][0] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat1roll":    arti.get("sec_effects", [[None]*5]*4)[0][1] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat1crafted": arti.get("sec_effects", [[None]*5]*4)[0][3] if len(arti.get("sec_effects", [])) > 0 and len(arti.get("sec_effects", [[None]*5]*4)[0]) > 2 else None,
                            #stat2
                            "stat2":        arti.get("sec_effects", [[None]*5]*4)[1][0] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat2roll":    arti.get("sec_effects", [[None]*5]*4)[1][1] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat2crafted": arti.get("sec_effects", [[None]*5]*4)[1][3] if len(arti.get("sec_effects", [])) > 1 and len(arti.get("sec_effects", [[None]*5]*4)[1]) > 2 else None,
                            #stat3
                            "stat3":        arti.get("sec_effects", [[None]*5]*4)[2][0] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat3roll":    arti.get("sec_effects", [[None]*5]*4)[2][1] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat3crafted": arti.get("sec_effects", [[None]*5]*4)[2][3] if len(arti.get("sec_effects", [])) > 2 and len(arti.get("sec_effects", [[None]*5]*4)[2]) > 2 else None,
                            #stat4
                            "stat4":        arti.get("sec_effects", [[None]*5]*4)[3][0] if arti.get("sec_effects", [[None]*5]*4)[1] else None,
                            "stat4roll":    arti.get("sec_effects", [[None]*5]*4)[3][1] if arti.get("sec_effects", [[None]*5]*4)[0] else None,
                            "stat4crafted": arti.get("sec_effects", [[None]*5]*4)[3][3] if len(arti.get("sec_effects", [])) > 3 and len(arti.get("sec_effects", [[None]*5]*4)[3]) > 2 else None,

                            "category": arti.get("locked"),
                            "extra": str(arti.get("extra")),
                            "date_add": arti.get("date_add"),
                            "date_mod": arti.get("date_mod"),
                        }
                        monster_rows.append(row)
            monster_df = pd.DataFrame(monster_rows)

            # statXroll-Spalten als float definieren
            for col in ["stat1roll", "stat2roll", "stat3roll", "stat4roll"]:
                if col in artifact_df.columns:
                    artifact_df[col] = pd.to_numeric(artifact_df[col], errors="coerce")
            for col in ["stat1roll", "stat2roll", "stat3roll", "stat4roll"]:
                if col in monster_df.columns:
                    monster_df[col] = pd.to_numeric(monster_df[col], errors="coerce")

            def attribute_type_mapper(row):
                try:
                    attribute_map = {
                        1: "Wasser",
                        2: "Feuer",
                        3: "Wind",
                        4: "Licht",
                        5: "Dunkel",
                        98: "Immateriell"
                    }
                    unit_style_map = {
                        1: "Attack",
                        2: "Defense",
                        3: "HP",
                        4: "Support",
                        5: "Material",
                        98: "Immateriell"
                    }
                    # Wenn attribute fehlt, 0 ist oder None, nutze unit_style
                    if pd.isnull(row["attribute"]) or int(row["attribute"]) == 0:
                        attribute_str = unit_style_map.get(int(row.get("unit_style", 0)), "Unbekannt")
                    else:
                        attribute_str = attribute_map.get(int(row["attribute"]), "Unbekannt")
                    return f"{attribute_str}"
                except Exception:
                    return "Unbekannt"

            artifact_df["attribute"] = artifact_df.apply(attribute_type_mapper, axis=1)
            monster_df["attribute"] = monster_df.apply(attribute_type_mapper, axis=1)
            # Entferne die Hilfsspalte unit_style nach Verarbeitung
            if "unit_style" in artifact_df.columns:
                artifact_df.drop(columns=["unit_style"], inplace=True)
            if "unit_style" in monster_df.columns:
                monster_df.drop(columns=["unit_style"], inplace=True)

            def type_mapper(x):
                try:
                    return "Rechts" if int(x) == 2 else "Links"
                except (ValueError, TypeError):
                    return "ERROR"

            def natural_rank_mapper(x):
                mapping = {
                    1: "Normal",
                    2: "Magic",
                    3: "Rare",
                    4: "Hero",
                    5: "Legendary"
                }
                return mapping.get(int(x), x) if pd.notnull(x) else x

            def pri_mapper(x):
                try:
                    if int(x[0]) == 100:
                        return "HP"
                    elif int(x[0]) == 101:
                        return "ATK"
                    return "DEF"
                except (ValueError, TypeError):
                    return "ERROR"

            def stat_mapper(x):
                try:
                    mapping = {
                        200: "ATK Increased Proportional to Lost HP up to x %",
                        201: "DEF Increased Proportional to Lost HP up to x %",
                        202: "SPD Increased Proportional to Lost HP up to x %",
                        203: "SPD Under Inability Effects",
                        204: "ATK Increasing Effect",
                        205: "DEF Increasing Effect",
                        206: "SPD Increasing Effect",
                        207: "Crit Rate Increasing Effect",
                        208: "Damage Dealt by Counterattack",
                        209: "Damage Dealt by Attacking Together",
                        210: "Bomb Damage",
                        211: "Damage Dealt by Reflected DMG",
                        212: "Crushing Hit DMG",
                        213: "Damage Received Under Inability Effect",
                        214: "Received Crit DMG",
                        215: "Life Drain",
                        216: "HP when Revived",
                        217: "Attack Bar when Revived",
                        218: "Additional Damage by % of HP",
                        219: "Additional Damage by % of ATK",
                        220: "Additional Damage by % of DEF",
                        221: "Additional Damage by % of SPD",
                        222: "CRIT DMG+ as the enemy's HP condition is good",
                        223: "CRIT DMG+ as the enemy's HP condition is bad",
                        224: "Single-target skill CRIT DMG on your turn",
                        225: "Counterattack/Co-op Attack DMG",
                        226: "ATK/DEF UP Effect",
                        300: "Damage Dealt on Fire",
                        301: "Damage Dealt on Water",
                        302: "Damage Dealt on Wind",
                        303: "Damage Dealt on Light",
                        304: "Damage Dealt on Dark",
                        305: "Damage Received from Fire",
                        306: "Damage Received from Water",
                        307: "Damage Received from Wind",
                        308: "Damage Received from Light",
                        309: "Damage Received from Dark",
                        400: "Skill 1 CRIT DMG",
                        401: "Skill 2 CRIT DMG",
                        402: "Skill 3 CRIT DMG",
                        403: "Skill 4 CRIT DMG",
                        404: "Skill 1 Recovery",
                        405: "Skill 2 Recovery",
                        406: "Skill 3 Recovery",
                        407: "Skill 1 Accuracy",
                        408: "Skill 2 Accuracy",
                        409: "Skill 3 Accuracy",
                        410: "[Skill 3/4] CRIT DMG",
                        411: "First Attack CRIT DMG"
                    }
                    return mapping.get(int(x), str(x))
                except Exception:
                    return

            def stat_crafted_mapper(x):
                try:
                    mapping = {
                        0: "Nein",
                        1: "Ja"
                    }
                    return mapping.get(int(x), str(x))
                except Exception:
                    return "ERROR"

            column_map = {
                "type": type_mapper,
                "natural_rank": natural_rank_mapper,
                "pri_effect": pri_mapper,
                "stat1": stat_mapper,
                "stat1crafted": stat_crafted_mapper,
                "stat2": stat_mapper,
                "stat2crafted": stat_crafted_mapper,
                "stat3": stat_mapper,
                "stat3crafted": stat_crafted_mapper,
                "stat4": stat_mapper,
                "stat4crafted": stat_crafted_mapper,
            }

            for col, func in column_map.items():
                if col in artifact_df.columns:
                    artifact_df[col] = artifact_df[col].apply(func)
            for col, func in column_map.items():
                if col in monster_df.columns:
                    monster_df[col] = monster_df[col].apply(func)

            # Excel speichern
            save_path = filedialog.asksaveasfilename(
                title="Speichere als Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel Datei", "*.xlsx"), ("Alle Dateien", "*.*")]
            )
            if save_path:
                with pd.ExcelWriter(save_path, engine="openpyxl") as writer:
                    artifact_df.to_excel(writer, sheet_name="Artefakte", index=False)
                    monster_df.to_excel(writer, sheet_name="Monster", index=False)
                # Nach dem Schreiben: Format und Spaltenbreite anpassen
                import openpyxl
                wb = openpyxl.load_workbook(save_path)
                # Definiere feste Spaltenbreiten
                custom_widths = {
                    "A": 12, "I": 12, "L": 12, "O": 12, "R": 12,
                    "B": 14, "E": 14,
                    "G": 50, "J": 50, "M": 50, "P": 50,
                    "U": 20, "V": 20
                }
                for sheet_name in ["Artefakte", "Monster"]:
                    ws = wb[sheet_name]
                    for col_letter, width in custom_widths.items():
                        ws.column_dimensions[col_letter].width = width
                    # Zahlenformat für rid und occupied_id
                    for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        header = ws.cell(row=1, column=col[0].column).value
                        if header in ["rid", "occupied_id"]:
                            for cell in col:
                                cell.number_format = '0'
                wb.save(save_path)
                messagebox.showinfo("Erfolg", f"Datei gespeichert unter:\n{save_path}")
        except Exception as e:
            messagebox.showerror("Fehler beim Export", str(e))

    root = tk.Tk()
    root.title("Datei auswählen")

    entry_var = tk.StringVar()

    entry = tk.Entry(root, textvariable=entry_var, width=50)
    entry.pack(padx=10, pady=10)

    browse_btn = tk.Button(root, text="Durchsuchen", command=browse_file)
    browse_btn.pack(padx=10, pady=5)

    export_btn = tk.Button(root, text="Artefakte exportieren", command=export_artifacts)
    export_btn.pack(padx=10, pady=5)

    root.mainloop()

if __name__ == "__main__":
    main()
